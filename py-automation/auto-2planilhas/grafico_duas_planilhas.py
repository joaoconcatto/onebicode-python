from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, Series

dict_anos = {}

# 1- importando despesas.xlsx
arquivo_despesas = load_workbook(filename='auto-2planilhas/planilhas/despesas.xlsx')
ws_despesas = arquivo_despesas['Sheet']
max_linhas = ws_despesas.max_row

for i in range(2, max_linhas+1):
    # print(ws_despesas['A%s' %i].value)
    # print(ws_despesas['B%s' %i].value)

    dict_anos[ws_despesas['A%s' %i].value] = {'Despesa': ws_despesas['B%s' %i].value, 'Receita': 0}


# 2- importando receitas.xlsx
arquivo_receitas = load_workbook(filename='auto-2planilhas/planilhas/receitas.xlsx')
ws_receitas = arquivo_receitas['Sheet']
max_linhas = ws_receitas.max_row

for i in range(2, max_linhas+1):
    dict_anos[ws_receitas['A%s' %i].value]['Receita'] = ws_receitas['B%s' %i].value


print(dict_anos)


# 3- criando a planilha
wb = Workbook()
ws = wb.active

ws['A1'] = 'Ano'
ws['B1'] = 'Despesas'
ws['C1'] = 'Receitas'

i = 2

for key, value in dict_anos.items():
    ws['A%s' %i] = key
    ws['B%s' %i] = value['Despesa']
    ws['C%s' %i] = value['Receita']

    i += 1

chart1 = BarChart()
chart1.type = 'col'
chart1.style = 12
chart1.title = 'Receita x Despesa'
chart1.x_axis.title = 'Ano'
chart1.y_axis.title = 'R$'

data = Reference(
    ws,
    min_col=2,
    max_col=3,
    min_row=1,
    max_row=i

)

anos = Reference(
    ws,
    min_col=1,
    min_row=2,
    max_row=i
)

chart1.add_data(data, titles_from_data=True)
chart1.set_categories(anos)
chart1.shape = 4
ws.add_chart(chart1, 'A%s' %(i+2))

wb.save(filename='auto-2planilhas/planilhas/demonstrativo.xlsx')