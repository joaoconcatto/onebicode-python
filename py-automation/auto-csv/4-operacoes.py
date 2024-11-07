from openpyxl import load_workbook
from openpyxl.drawing.image import Image


wb = load_workbook(filename='auto-csv/planilhas/gastos.xlsx')
planilha = wb['Planilha1']


valor_total = 0

# 1- somando os valores de uma planilha
for i in range(2, 8):
    valor = int(planilha['B%s' %i].value)
    valor_total += valor
    
print(valor_total)


# 2- adicionar em uma célula da planilha o valor da operação
planilha['B8'] = valor_total

# wb.save(filename='auto-csv/planilhas/gastos_totalizado.xlsx')


# 3- mesclar células
planilha['A11'] = 'Teste'
planilha.merge_cells('A11:B11')
# wb.save(filename='auto-csv/planilhas/gastos_totalizado.xlsx')


# 4- inserir imagens
img  = Image('auto-email/bb_preco.png')
planilha.add_image(img, 'C11')
wb.save(filename='auto-csv/planilhas/gastos_imagem.xlsx')