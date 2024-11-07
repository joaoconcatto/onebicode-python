from openpyxl import load_workbook, Workbook

lista_arquivos = ['gastos_imagem', 'gastos_totalizado', 'gastos', 'planilha1']

# 1- criar nova planilha
wb = Workbook()
nome_arquivo = 'auto-csv/planilhas/resultado.xlsx'

for nome in lista_arquivos:
    arquivo = load_workbook(filename='auto-csv/planilhas/%s.xlsx' %nome)
    sheet = arquivo[nome]
    max_linhas = sheet.max_row
    max_colunas = sheet.max_column

    ws = wb.create_sheet(title=nome)

    # 2- iterar valores da planilha
    for i in range(1, max_linhas+1):
        for j in range(1, max_colunas+1):
            data = sheet.cell(row=i, column=j)
            ws.cell(row=i, column=j).value = data.value

wb.save(nome_arquivo)